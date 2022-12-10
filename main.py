import requests

import re
import math
import openpyxl
import time
import numpy
import getpass
import json
import csv
import datetime
import calendar
import argparse
global PRTG_PASSWORD

def xlsx_build(): ### BUILDS XLSX FILE
    now_datetime = datetime.datetime.now()
    now = now_datetime.date()

    t0_headers = now
    t14_headers = now - datetime.timedelta(days = 14)
    t28_headers = now - datetime.timedelta(days = 28)

    t0_headers = t0_headers.strftime('%m/%d')
    t14_headers = t14_headers.strftime('%m/%d')
    t28_headers = t28_headers.strftime('%m/%d')

    sheetHeaders = ['Location','Max Traffic (Mb/s)','Choke Point (Device)','Choke Point Limit (Mb/s)','Circuit Max Limit (Mb/s)',
        'Circuit Utilization', f'Choke Utilization (Current)',
        f'Choke Utilization ({t14_headers})',
        f'Choke Utilization ({t28_headers})',
        'Max Usage Plan','Notes','Action']

    coreUtilSummaryHeaders = ['Core Utilization Summary','Bandwidth (Mb/s)', 'Max Capacity',
        f'Gross Utilization (Current)',
        f'Gross Utilization ({t14_headers})',
        f'Gross Utilization ({t28_headers})']

    global alphabetArray
    alphabetArray = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

    for letter in alphabetArray:
        outputMainSheet.column_dimensions[str(letter)].width = '16'
        outputSummarySheet.column_dimensions[str(letter)].width = '16'

    outputMainSheet.column_dimensions['A'].width = '28'
    outputSummarySheet.column_dimensions['A'].width = '28'
    outputMainSheet.column_dimensions['K'].width = '12'
    outputSummarySheet.column_dimensions['K'].width = '12'
    outputMainSheet.column_dimensions['F'].width = '12'
    outputSummarySheet.column_dimensions['F'].width = '16'
    outputMainSheet.column_dimensions['L'].width = '14'
    outputSummarySheet.column_dimensions['L'].width = '14'
    outputMainSheet.column_dimensions['M'].width = '14'
    outputSummarySheet.column_dimensions['M'].width = '14'
    outputMainSheet.column_dimensions['B'].width = '12'
    outputSummarySheet.column_dimensions['B'].width = '12'

    #        if sensorTagData.get('edge'):
    for j in range(1,500):
        outputMainSheet.cell(row=int(j),column=int(6)).style='Percent'
        outputMainSheet.cell(row=int(j),column=int(7)).style='Percent'
        outputMainSheet.cell(row=int(j),column=int(8)).style='Percent'
        outputMainSheet.cell(row=int(j),column=int(9)).style='Percent'
        outputSummarySheet.cell(row=int(j),column=int(4)).style='Percent'
        outputSummarySheet.cell(row=int(j),column=int(5)).style='Percent'
        outputSummarySheet.cell(row=int(j),column=int(6)).style='Percent'

    for i in range(0,len(sheetHeaders)):
        outputMainSheet[str(alphabetArray[i])+'1']=sheetHeaders[i]
        outputMainSheet[str(alphabetArray[i])+'1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)


    for i in range(0,len(coreUtilSummaryHeaders)):
        outputSummarySheet[str(alphabetArray[i])+'1']=coreUtilSummaryHeaders[i]
        outputSummarySheet[str(alphabetArray[i])+'1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    outputSummarySheet['A5']='Total: '

    #outputMainSheet['B2:M300'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    outputWorkbook.save("hopefullythisworks.xlsx")



PRTG_PASSWORD = "M9y%23asABUx9svvs"  ###### !! CHANGE FOR PROD !! ##### ----------
PRTG_HOSTNAME = 'nanm.smartaira.net'   ###!! Static, domain/URL to PRTG server
#PRTG_HOSTNAME = '64.71.154.163'
#PRTG_PASSWORD = getpass.getpass('Password: ')

outputWorkbook = openpyxl.Workbook()

sheet1_to_be_del = outputWorkbook.get_sheet_by_name('Sheet')
outputWorkbook.remove_sheet(sheet1_to_be_del)

outputMainSheet = outputWorkbook.create_sheet("Property Bandwidths")
outputSummarySheet = outputWorkbook.create_sheet("Summaries")

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import colors

alertRule = ColorScaleRule(start_type='min', start_value=0, start_color=colors.WHITE, end_type='max', 
    end_value=100, end_color='F8696B')

outputMainSheet.conditional_formatting.add("F2:J250", alertRule)
outputSummarySheet.conditional_formatting.add("C2:F40", alertRule)
## [disk io] 
xlsx_build() # Creates XLSX file and defines headers, column widths, row heights, etc. 
                # Saves to 'hopefullythisworks.xlsx', needs to be changed to reflect proper naming scheme
print("XLSX file template built successfully")

outputWorkbook.save("hopefullythisworks.xlsx")



now = datetime.datetime.now()
default_start = now - datetime.timedelta(days = 14)
default_end = now - datetime.timedelta(days = 1)

back_14 = now - datetime.timedelta(days = 14)

back_28 = now - datetime.timedelta(days = 28)



parser = argparse.ArgumentParser()
parser.add_argument('--username', default='agriffin', required=False,
                  help='PRTG username for API call')
parser.add_argument('--start', default=default_start.strftime('%Y-%m-%d'),
                  help='Historic data start date (yyyy-mm-dd)')
parser.add_argument('--end', default=default_end.strftime('%Y-%m-%d'),
                  help='Historic data end date (yyyy-mm-dd)')
parser.add_argument('--avgint', default="3600",
                  help='Averaging interval. Smaller numbers increase api call time!'
                       ' Default is 3600')
parser.add_argument('--debug', action='store_true', dest='debug',
                  help='add additional debugging fields to output')
parser.add_argument('--percentile', default="99",
                  help='set the percentile for reporting (default is 90)')
parser.add_argument('--output',
                  help='specify output file directory (default is cwd)')
parser.add_argument('--sensorid',
                  help='Pull data from only one sensorid')
args = parser.parse_args()



# Prompt the user for credentials
#PRTG_PASSWORD = getpass.getpass('Password: ')

def get_kpi_sensor_ids(username, password):
    """
    Returns a dict with the following format:
    {'prtg-version': '22.1.74.1869',
    'treesize': 2,
    'sensors': [{'objid': 14398,
      'objid_raw': 14398,
      'device': 'ACA Edge (160.3.214.2)',
      'device_raw': 'ACA Edge (160.3.214.2)'},
    """
    response = requests.get(
            f'https://{PRTG_HOSTNAME}/api/table.json?content=sensors&output=json'
            f'&columns=objid,device,tags&filter_tags=kpi_bandwidth'
            f'&username={username}&password={password}&sortby=device')
    if response.status_code == 200:
        response_tree = json.loads(response.text)
        if args.sensorid:
            for i in response_tree.get('sensors'):
                if i['objid'] == int(args.sensorid):
                    return [i]
        else:
            return response_tree.get('sensors')
    else:
        print("Error in sensor list request")
        exit(1)



def normalize_traffic(data, label):
    """
    Takes an input of raw PRTG historic data and the label (ie 'Traffic In (speed)')
    and multiplies the speeds by 0.00008 to get values in mbits/sec ( ((8 / 10) / 100) / 1000)
    """
    traffic_list = []
    for i in data['histdata']:
        try:
            if i[str(label)] !='':
                traffic_list.append(int(i[label]) * 0.000008)
        except:
            traffic_list.append(1)
        else:
            traffic_list.append(1)
    return traffic_list

def extract_tags(sensor):
    """
    Takes a sensor dictionary and extacts a properties dictionary from the tags
    string returned by PRTG. The tags string will have a format something like this:
    'kpi_bandwidth kpi_seg=DIA kpi_choke=Circuit kpi_chokelimit=10000 kpi_cktmaxlimit=10000'
    
    """
    target_tags = ['kpi_seg', 'kpi_choke', 'kpi_cktmaxlimit', 'kpi_siteid', 'edge']
    def filter_tags(tags_list, property_string):
        """
        Takes a list of tags and splits them into a properties dict:
        ['kpi_choke=Circuit', kpi_chokelimit=10000'] becomes
        {'kpi_choke':'Circuit', 'kpi_chokelimit':'10000'}
        
        This function exists because we might have tags like kpi_choke and kpi_chokelimit
        that will both be found by the _in_ function
        """
        properties = {}
        property_list = list(filter(lambda a: property_string in a, tags_list))
        for property in property_list:
            try:
                key, value = property.split('=')
                properties.update({key:value})
            except:
                properties["edge"] = "True"
        return properties

    device_properties = {}
    tag_string = sensor['tags'].split()

    for tag in target_tags:
        device_properties.update(filter_tags(tag_string, tag))

    return device_properties

summary_data = []

sensors = get_kpi_sensor_ids(args.username, PRTG_PASSWORD)
# ! # ! # ! output_file = f'output_{args.start}--{args.end}.csv'
#if args.output:
#    if args.output.endswith('/'):
#        output_file = args.output + output_file
#    else:
#        output_file = args.output + '/' + output_file
#if args.debug:
#    headers.insert(1, 'Device id')
#    headers.insert(1, 'Device')




def write_XLSX(row, col, val):
    outputMainSheet.cell(row=int(row),column=int(col)).value=val
    outputWorkbook.save("hopefullythisworks.xlsx")

def write_XLSX_seg(row, col, val):
    outputSummarySheet.cell(row=int(row),column=int(col)).value=val
    outputWorkbook.save("hopefullythisworks.xlsx")


row_index = 2
for sensor in sensors:
    response = requests.get(
            f'https://{PRTG_HOSTNAME}/api/historicdata.json?id={sensor["objid"]}'
            f'&avg={args.avgint}&sdate={args.start}-00-00&edate={args.end}-23-59'
            f'&usecaption=1'
            f'&username={args.username}&password={PRTG_PASSWORD}')
    if response.status_code == 200:

        data = json.loads(response.text)

        properties = extract_tags(sensor)

        traffic_in = normalize_traffic(data, 'Traffic In (speed)')

        traffic_out = normalize_traffic(data, 'Traffic Out (speed)')
        
        device_name = sensor['device'].split(' (')[0]

        out_array = []

        # LOCATION
        if properties.get('kpi_siteid'):
            out_array.append(re.sub("#", " ", properties['kpi_siteid']))
        else:
            out_array.append('NA')

#        if args.debug:
            # Device name (debug)
#            out_array.append(device_name)

            # Device id (debug)
#            out_array.append(sensor['objid'])

        # MAX TRAFFIC
        max_traffic = 0
        if properties.get('kpi_trafficdirection') == 'up':
            if traffic_out == []:
                out_array.append("NA")
            else:
                max_traffic = math.ceil(numpy.percentile(traffic_out, int(args.percentile)))
                out_array.append(max_traffic)
        else:
            if traffic_in == []:
                out_array.append("NA")
            else:
                max_traffic = math.ceil(numpy.percentile(traffic_in, int(args.percentile)))
                out_array.append(max_traffic)

        if "Core" in properties.get('kpi_seg'):
            summary_data.append({'segment': properties.get('kpi_seg'),
                                 'bandwidth': max_traffic,
                                 'limit': properties.get('kpi_cktmaxlimit')})
        else:
            pass

        # CHOKE POINT DEVICE 'kpi_choke'
        if properties.get('kpi_choke'):
            out_array.append(properties['kpi_choke'])
        else:
            out_array.append('NA')

        # CHOKE POINT LIMIT 'kpi_chokelimit'
        if properties.get('kpi_chokelimit'):
            out_array.append(int(properties['kpi_chokelimit']))
        else:
            out_array.append('NA')

        # CIRCUIT MAX LIMIT 'kpi_cktmaxlimit'
        if properties.get('kpi_cktmaxlimit'):
            out_array.append(int(properties['kpi_cktmaxlimit']))
        else:
            out_array.append('NA')

        # CIRCUIT UTILIZATION
        # max down / circuit max limit
        if properties.get('kpi_cktmaxlimit'):
            out_array.append(max_traffic / int(properties['kpi_cktmaxlimit']))
        else:
            out_array.append('NA')

        # CHOKE UTILIZATION CURRENT
        # max down / choke point limit
        if properties.get('kpi_chokelimit'):
            out_array.append(int(max_traffic) / int(properties['kpi_chokelimit']))
        else:
            out_array.append('NA')  

        ### EDGE BOOLEAN
        if properties.get('edge'):
            outputMainSheet.cell(row=int(row_index),column=int(1)).fill=openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
            outputMainSheet.cell(row=int(row_index),column=int(2)).fill=openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
            outputMainSheet.cell(row=int(row_index),column=int(3)).fill=openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
            outputMainSheet.cell(row=int(row_index),column=int(4)).fill=openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
            outputMainSheet.cell(row=int(row_index),column=int(5)).fill=openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
        else:
            pass

        i=1
        while i <= len(out_array):
            write_XLSX(row_index, int(i), out_array[int(i-1)])
            i += 1
        row_index += 1
    else: 
        print("Error occurred")





segments = set()
for data in summary_data:
    # Create set from all the segment values (creates a unique list)
    segments.add(data.get('segment'))

segment_bandwidth_total = 0
segment_capacity_total = 0

seg_row_index = 2
for segment in segments:
    segment_bandwidth = 0
    segment_limit = 0
    for data in summary_data:
        if data['segment'] == segment:
            segment_bandwidth += int(data['bandwidth'])
            segment_limit += int(data['limit'])
    saturation = segment_bandwidth / segment_limit
    segment_bandwidth_total += segment_bandwidth
    segment_capacity_total += segment_limit

    segment_list = [segment, segment_bandwidth, segment_limit, saturation]

    seg_col_index = 1
    while seg_col_index < len(segment_list):
        write_XLSX_seg(seg_row_index, seg_col_index, segment_list[int(seg_col_index)-1])
        seg_col_index += 1
    seg_row_index += 1


segment_saturation = segment_bandwidth_total / segment_capacity_total

segment_total_list=[segment_bandwidth_total, segment_capacity_total, segment_saturation]


seg_tot_row_index = 5
seg_tot_col_index = 2
while seg_tot_col_index < (len(segment_total_list)+2):
    write_XLSX_seg(seg_tot_row_index, seg_tot_col_index, segment_list[int(seg_tot_col_index)-2])
    seg_tot_col_index += 1