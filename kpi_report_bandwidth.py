# --------   PRTG-XLSX-Report-Generator.py
# -------------------------------------------------------------------------------
#
#   Pulls sensor & device data from PRTG API and neatly formats the data into a .xmlx (MS Excel) file
#       using python-openpyxl, json, csv, and pandas. 
#   
#       Average runtime: 274 seconds (4.5 - 5.0 minutes)
# -------------------------------------------------------------------------------
from time import time, time_ns
from turtle import width
import requests
import re
import math
import numpy
import getpass
import json

import datetime
import argparse
import openpyxl
import os
import pprint

###########################################################
### [Most primary functions declared here]

### [Getting script's current working directory to be used later to ensure that 
#       no exceptions arise and for safer file IO]
#############


### [Time-Frames/Time Windows For Pulling Historical Data from PRTG]
#############
def timeWindowFrames(timeFrameIDRAW):
    timeFrameID = str(timeFrameIDRAW)
    #############
    ### [Time window declarations]
    ### [A positive time in these comments indicates # of days prior to current (DAY 0)]
    ### [eg: "0d -- 14d" = "From today (0d) through 14 days ago (14d)"]
    #############
    current_sys_datetime = datetime.datetime.now()
    ######    0d -- 14d
    def_s = current_sys_datetime - datetime.timedelta(days = 14)
    def_e = current_sys_datetime - datetime.timedelta(days = 1)
    ######    7d -- 21d
    win1_s = current_sys_datetime - datetime.timedelta(days = 21)
    win1_e = current_sys_datetime - datetime.timedelta(days = 7)
    ######    14d -- 28d
    win2_s = current_sys_datetime - datetime.timedelta(days = 28)
    win2_e = current_sys_datetime - datetime.timedelta(days = 14)
    ######    21d -- 35d
    win3_s = current_sys_datetime - datetime.timedelta(days = 35)
    win3_e = current_sys_datetime - datetime.timedelta(days = 21)

    # return [def_s,def_e],[win1_s,win1_e],[win2_s,win2_e],[win3_s,win3_e]
    if timeFrameID == "0":
        return def_s,def_e
    elif timeFrameID == "1":
        return win1_s,win1_e
    elif timeFrameID == "2":
        return win2_s,win2_e
    elif timeFrameID == "3":
        return win3_s,win3_e
    else: 
        print("Error: timeWindowFrames -- Invalid arguments passed")
### [Defining path to Temporary file]
#############

### [CLI argument parser; --username is required]
#############
def cliArgumentParser(currentSystemDatetime):
    now = datetime.datetime.now()
    default_start = now - datetime.timedelta(days = 28)
    default_end = now - datetime.timedelta(days = 0)
    parser = argparse.ArgumentParser()
    parser.add_argument('--username', required=False, default="agriffin", ###### !! CHANGE FOR PROD !! #####
                    help='PRTG username for API call')
    parser.add_argument('--start', default=default_start.strftime('%Y-%m-%d'),
                    help='Historic data start date (yyyy-mm-dd)')
    parser.add_argument('--end', default=default_end.strftime('%Y-%m-%d'),
                    help='Historic data end date (yyyy-mm-dd)')
    parser.add_argument('--avgint', default="21600",
                    help='Averaging interval. Smaller numbers increase api call time!'
                        ' Default is 21600 seconds (6 hours)')
    parser.add_argument('--debug', action='store_true', dest='debug',
                    help='add additional debugging fields to output')
    parser.add_argument('--percentile', default="99",
                    help='set the percentile for reporting (default is 90)')
    parser.add_argument('--output',
                    help='specify output file directory (default is cwd)')
    parser.add_argument('--sensorid',
                    help='Pull data from only one sensorid')
    cliargs = parser.parse_args()
    return cliargs

def xlsx_build():
    t0BACK_headers_u,t14BACK_headers_u = timeWindowFrames("0")
    t7BACK_headers_u,t21BACK_headers_u = timeWindowFrames("1")
    t14BACK_headers_u,t28BACK_headers_u = timeWindowFrames("2")
    t21BACK_headers_u,t35BACK_headers_u = timeWindowFrames("3")

    t0BACK_headers = t0BACK_headers_u.strftime('%m/%d')
    t7BACK_headers = t7BACK_headers_u.strftime('%m/%d')
    t14BACK_headers = t14BACK_headers_u.strftime('%m/%d')
    t21BACK_headers = t21BACK_headers_u.strftime('%m/%d')
    t28BACK_headers = t28BACK_headers_u.strftime('%m/%d')
    t35BACK_headers = t35BACK_headers_u.strftime('%m/%d')


    sheetHeaders = ['Location','Highest Traffic (Mb/s)','Choke Point (Device)','Choke Point Throttle (Mb/s)','Circuit Max Limit (Mb/s)',
        'Circuit Utilization',
        f'Choke Utilization ({t14BACK_headers} - {t0BACK_headers})',
        f'Choke Utilization ({t21BACK_headers} -  {t7BACK_headers})',
        f'Choke Utilization ({t28BACK_headers} - {t14BACK_headers})',
        f'Choke Utilization ({t35BACK_headers} - {t21BACK_headers})',
        'Max Usage Plan','Notes','Action']

    coreUtilSummaryHeaders = ['Core Utilization Summary','Bandwidth (Mb/s)',
        f'Gross Utilization ({t14BACK_headers} - {t0BACK_headers})',
        f'Gross Utilization ({t21BACK_headers} - {t7BACK_headers})',
        f'Gross Utilization ({t28BACK_headers} - {t14BACK_headers})']

    alphabetArray = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

    for letter in alphabetArray:
        outputMainSheet.column_dimensions[str(letter)].width = '16'
    outputMainSheet.column_dimensions['A'].width = '23'
    outputMainSheet.column_dimensions['K'].width = '12'
    outputMainSheet.column_dimensions['F'].width = '12'
    outputMainSheet.column_dimensions['L'].width = '14'
    outputMainSheet.column_dimensions['M'].width = '14'
    outputMainSheet.column_dimensions['B'].width = '12'

    for i in range(0,len(sheetHeaders)):
        outputMainSheet[str(alphabetArray[i])+'7']=sheetHeaders[i]
        outputMainSheet[str(alphabetArray[i])+'7'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)


    for i in range(0,len(coreUtilSummaryHeaders)):
        outputMainSheet[str(alphabetArray[i])+'1']=coreUtilSummaryHeaders[i]
        outputMainSheet[str(alphabetArray[i])+'1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    outputMainSheet['A5']='Total: '

    outputWorkbook.save("test.xlsx")

def writeToSheet(yIndex,xIndex,vIndex):
    outputMainSheet.cell(row=int(yIndex),column=int(xIndex)).value=vIndex
    outputWorkbook.save("test.xlsx")

def get_kpi_sensor_ids(username, password, PRTG_HOSTNAME):
    """
    Returns a dict with the following format:
    {'prtg-version': '22.1.74.1869',
    'treesize': 2,
    'sensors': [{'objid': 14398,
    'objid_raw': 14398,
    'device': 'ACA Edge (160.3.214.2)',
    'device_raw': 'ACA Edge (160.3.214.2)'},
    """
   
    response = requests.get(
            f'https://{PRTG_HOSTNAME}/api/table.json?content=sensors&output=json'
            f'&columns=objid,device,tags&filter_tags=kpi_bandwidth'
            f'&username={username}&password={password}&sortby=device', verify=False
            )
    ### If 200 OK HTTP response is not seen, raise error and print cause to terminal
    if response.status_code == 200:
        response_tree = json.loads(response.text)
        print("Sensor data successfully queried from the API")
        if cliargs.sensorid:
            for i in response_tree.get('sensors'):
                if i['objid'] == int(cliargs.sensorid):
                    print(i)
                    return [i]
                    
        else:

            return response_tree.get('sensors')
    else:
        print("Error making API call to nanm.smartaira.net")
        print("HTTP response 200: OK was not received")
        print("Received response code: "+str(response.status_code))
        quit()


def normalize_traffic(data, label):
    """
    Takes an input of raw PRTG historic data and the label (ie 'Traffic In (speed)')
    and multiplies the speeds by 0.00008 to get values in mbits/sec ( ((8 / 10) / 100) / 1000)
    """
    traffic_list = []
    for i in data['histdata']:
        if i[label] != '':
            traffic_list.append(i[label] * 0.000008)
    return traffic_list

def extract_tags(sensor):
    """
    Takes a sensor dictionary and extacts a properties dictionary from the tags
    string returned by PRTG. The tags string will have a format something like this:
    'kpi_bandwidth kpi_seg=DIA kpi_choke=Circuit kpi_chokelimit=10000 kpi_cktmaxlimit=10000'
    
    """
    target_tags = ['kpi_seg', 'kpi_choke', 'kpi_cktmaxlimit', 'kpi_siteid']
    def filter_tags(tags_list, property_string):
        """
        Takes a list of tags and splits them into a properties dict:
        ['kpi_choke=Circuit', kpi_chokelimit=10000'] becomes
        {'kpi_choke':'Circuit', 'kpi_chokelimit':'10000'}
        
        This function exists because we might have tags like kpi_choke and kpi_chokelimit
        that will both be found by the _in_ function
        """
        properties = {}
        property_list = list(filter(lambda a: property_string in a, tags_list))
        for property in property_list:
            key, value = property.split('=')
            properties.update({key:value})
        return properties

    device_properties = {}
    tag_string = sensor['tags'].split()

    for tag in target_tags:
        device_properties.update(filter_tags(tag_string, tag))

    return device_properties

def extract_datetime(response_tree):
    timeFrame1_arr = []
    timeFrame1_arr.clear()
    timeFrame2_arr = []
    timeFrame2_arr.clear()
    timeFrame3_arr = []
    timeFrame3_arr.clear()
    timeFrame4_arr = []
    timeFrame4_arr.clear()

    now = datetime.datetime.now()
    now = now.date()

    historicResponseData = json.loads(response_tree.text)

    i = 0
    while i < len(historicResponseData['histdata']):
        


        i += 1
    #return timeFrame1_arr,timeFrame2_arr,timeFrame3_arr,timeFrame4_arr

    timeframe1_dict = {}
    timeframe1_dict["histdata"] = timeFrame1_arr
    timeframe2_dict = {}
    timeframe2_dict["histdata"] = timeFrame2_arr
    timeframe3_dict = {}
    timeframe3_dict["histdata"] = timeFrame3_arr
    timeframe4_dict = {}
    timeframe4_dict["histdata"] = timeFrame4_arr

    return [[timeframe1_dict,timeframe2_dict],[timeframe3_dict,timeframe4_dict]]

def buildComps(FrameWindow,datablock,historicResponseData,sensor,sensor_index):

    if 



    prtgDataDict['timeframes']['0 Weeks Back'] = {
      "FrameWindow": "0 Weeks Back",
      "treesize": 0,
      "histdata": [
    bbbb{
      "datetime": "",
      "Traffic Total (volume)": 0,
      "Traffic Total (speed)": 0,
      "Traffic In (volume)": 0,
      "Traffic In (speed)": 0,
      "Traffic Out (volume)": 0,
      "Traffic Out (speed)": 0,
      "Errors in (volume)": 0,
      "Errors in (speed)": 0,
      "Errors out (volume)": 0,
      "Errors out (speed)": 0,
      "Discards in (volume)": 0,
      "Discards in (speed)": 0,
      "Discards out (volume)": 0,
      "Discards out (speed)": 0,
      "Downtime": 0,
      "coverage": "0 %"
    }]}





    prtgDataDict[str(sensor['device'])]['historical']['data'] = {}

def storeAPIResponse(sensordata,historicResponseData,sensor,sensor_index):
    #    prtgDataDict[str(sensor['device'])]['objid'] = sensor['objid']
    i = 0
    while i < int(historicResponseData['treesize']):
        #dateBasedSectioning(sensordata, historicResponseData, sensor, sensor_index)
        indexHistoryDictionary = historicResponseData['histdata'][i]
        prtg_formatted_datetime = indexHistoryDictionary['datetime']
        try:
            regex_datetimesrc = re.search(r'\d{1}/\d{2}/\d{4}',prtg_formatted_datetime)
        except AttributeError:
            pass 
        if regex_datetimesrc:
            py_formatted_datetime = datetime.datetime.strptime(regex_datetimesrc.group(), '%m/%d/%Y')
        else: 
            try:
                regex_datetimesrc = re.search(r'\d{2}/\d{2}/\d{4}',prtg_formatted_datetime)
            except AttributeError:
                print("No date found in 'datetime' object!")
                pass
            if regex_datetimesrc:
                py_formatted_datetime = datetime.datetime.strptime(regex_datetimesrc.group(), '%m/%d/%Y')

        py_formatted_date = py_formatted_datetime.date()

        if py_formatted_date > now:
                print("Received PRTG data has out of range timestamp! (In the future)")
        elif py_formatted_date < (now - datetime.timedelta(days = 28)):
            print("Received PRTG data has out of range timestamp! (Too far back)")

        elif py_formatted_date > (now - datetime.timedelta(days = 7)):
            buildComps(0,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        elif py_formatted_date < (now - datetime.timedelta(days = 21)):
            buildComps(1,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        elif py_formatted_date > (now - datetime.timedelta(days = 21)) and py_formatted_date < (now - datetime.timedelta(days = 14)):
            buildComps(2,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        elif py_formatted_date < (now - datetime.timedelta(days = 7)) and py_formatted_date > (now - datetime.timedelta(days = 14)):
            buildComps(3,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        else: 
            print("Error extracting datetime from PRTG data!")


        i += 1
    '''
    for key,value in sensor.items():
        prtgDataDict[str(sensor['tags'])] = {}
        sensor_tags = sensor['tags']
        prtgDataDict[str(sensor['tags'])]['bandwidthsensor kpi_bandwidth kpi_choke=Circuit kpi_chokelimit=1000 kpi_cktmaxlimit=1000 kpi_seg=DIA kpi_siteid=94th#Aero router snmptrafficsensor router']
    '''

def prtgExtendHistParse1(timeFrame2_dict,sensor,i):

    timeFrame2_response = json.dumps(timeFrame2_dict)

    data = json.loads(timeFrame2_response.text)
    properties = extract_tags(sensor)
    traffic_in = normalize_traffic(data, 'Traffic In (speed)')
    traffic_out = normalize_traffic(data, 'Traffic Out (speed)')
    device_name = sensor['device'].split(' (')[0]

    ### [dec] - [MAX TRAFFIC (Mb/s)]
    #############
    max_traffic = 0
    if properties.get('kpi_trafficdirection') == 'up':
        if traffic_out == []:
            max_traffic_san = 0
        else:
            max_traffic = math.ceil(numpy.percentile(traffic_out, int(cliargs.percentile)))
            max_traffic_san = max_traffic
    else:
        if traffic_in == []:
            max_traffic_san = 0
        else:
            max_traffic = math.ceil(numpy.percentile(traffic_in, int(cliargs.percentile)))
            max_traffic_san = max_traffic

    ### [dec] - [CHOKE POINT UTILIZATION (%)]

    if properties.get('kpi_chokelimit'):
        outputMainSheet.cell(row=int(i),column=8).value=(float(max_traffic) / float(properties['kpi_chokelimit']))
        outputMainSheet.cell(row=int(i),column=8).style='Percent'
    else:
        outputMainSheet.cell(row=int(i),column=8).value='NA'

def prtgMainParse(timeFrame1_dict,sensor,kpi_seg_arr,s_count,i):
    
     # Declares 8th row on xlsx sheet to write mainResponseData to

    timeFrame1_response = json.dumps(timeFrame1_dict, indent=4)

    mainResponseData = json.loads(timeFrame1_response.text)
    sensorTagData = extract_tags(sensor)
    trafficInbound = normalize_traffic(mainResponseData, 'Traffic In (speed)')
    trafficOutbound = normalize_traffic(mainResponseData, 'Traffic Out (speed)')
    deviceName = sensor['device'].split(' (')[0]
    print("API data parsed and stored into memory")

    ### [dec] - [LOCATION (Location)]
    #############
    if sensorTagData.get('kpi_siteid'):
        writeToSheet(i,1,re.sub("#", " ", sensorTagData['kpi_siteid']))
    else:
        writeToSheet(i,1,'NA')

    if cliargs.debug:
        # Device name (debug)
        print(deviceName)
        # Device id (debug)
        print(sensor['objid'])

    ### [dec] - [MAX TRAFFIC (Mb/s)]
    #############
    max_traffic = 0
    if sensorTagData.get('kpi_trafficdirection') == 'up':
        if trafficOutbound == []:
            writeToSheet(i,2,'NA')
        else:
            max_traffic = math.ceil(numpy.percentile(trafficOutbound, int(cliargs.percentile)))
            writeToSheet(i,2,max_traffic)
    else:
        if trafficInbound == []:
            writeToSheet(i,2,'NA')
        else:
            max_traffic = math.ceil(numpy.percentile(trafficInbound, int(cliargs.percentile)))
            writeToSheet(i,2,max_traffic)

    ### [dec] - [CHOKE POINT (Device)]
    #############
    if sensorTagData.get('kpi_choke'):
        writeToSheet(i,3,sensorTagData['kpi_choke'])
    else:
        outputMainSheet.cell(row=int(i),column=3).value='NA'
        writeToSheet(i,3,'NA')

    ### [dec] - [CHOKE POINT LIMIT (Mb/s)]
    #############
    if sensorTagData.get('kpi_chokelimit'):
        writeToSheet(i,4,float(sensorTagData['kpi_chokelimit']))
        outputMainSheet.cell(row=int(i),column=int(4)).style='Percent'
    else:
        writeToSheet(i,4,'NA')
    


    ### [dec] - [CIRCUIT MAX LIMIT (Mb/s)]
    #############
    if sensorTagData.get('kpi_cktmaxlimit'):
        writeToSheet(i,5,float(sensorTagData['kpi_cktmaxlimit']))
    else:
        writeToSheet(i,5,'NA')

    ### [dec] - [CIRCUIT UTILIZATION (%)]
    #############
    if sensorTagData.get('kpi_cktmaxlimit'):
        writeToSheet(i,6,(float(max_traffic) / float(sensorTagData['kpi_cktmaxlimit'])))
        outputMainSheet.cell(row=int(i),column=int(6)).style='Percent'
    else:
        writeToSheet(i,6,'NA')


    ### [dec] - [CHOKE POINT UTILIZATION (%) 1]
    #############
    if sensorTagData.get('kpi_chokelimit'):
        writeToSheet(i,7,(float(max_traffic) / float(sensorTagData['kpi_chokelimit'])))
        outputMainSheet.cell(row=int(i),column=int(7)).style='Percent'
    else:
        writeToSheet(i,7,'NA') 


    if "Core" in sensorTagData.get('kpi_seg'):
        print("Core found, data will be written at the top of XLSX output")
        if sensorTagData.get('kpi_seg') in kpi_seg_arr:
            pass
        else:
            writeToSheet(1+s_count,1,sensorTagData.get('kpi_seg'))
            writeToSheet(1+s_count,2,max_traffic)
            writeToSheet(1+s_count,3,(float(max_traffic)/float(sensorTagData.get('kpi_cktmaxlimit'))))
            outputMainSheet.cell(row=int(1+s_count),column=3).style='Percent'
            s_count += 1
            kpi_seg_arr.append(sensorTagData.get('kpi_seg'))

def prtgMainCall(sensordata,PRTG_HOSTNAME,PRTG_PASSWORD,cliargs,kpi_seg_arr,s_count):
    i = 8
    sensor_index = 0
    print("Writing API query results to cache")
    for sensor in sensordata:
        print("Starting up new HTTPS session")
        response = requests.get(
            f'https://{PRTG_HOSTNAME}/api/historicdata.json?id={sensor["objid"]}'
            f'&avg={cliargs.avgint}&sdate={cliargs.start}-00-00&edate={cliargs.end}-23-59'
            f'&usecaption=1'
            f'&username={cliargs.username}&password={PRTG_PASSWORD}', verify=False
            )
        if response.status_code == 200:
            print("Initial PRTG API query successful")
            response_j = response.json()

            storeAPIResponse(sensordata,response_j,sensor,sensor_index)
            buildComps(sensordata,historicResponseData,sensor,sensor_index)

        else:
            print("Error making 'main' API call to nanm.smartaira.net (PRTG)")
            print("HTTP response 200: OK was not received")
            print("Received response code: "+str(response.status_code))
            exit(1)

        i += 1
        print("Closing up the HTTPS session")
        sensor_index += 1
### [PRTG API call and assigning data to "sensors" var]
#############


if __name__ == '__main__':
    
    PRTG_PASSWORD = "M9y%23asABUx9svvs"  ###### !! CHANGE FOR PROD !! ##### ----------
    PRTG_HOSTNAME = 'nanm.smartaira.net'   ###!! Static, domain/URL to PRTG server
    #PRTG_PASSWORD = getpass.getpass('Password: ')

    outputWorkbook = openpyxl.Workbook()
    outputMainSheet = outputWorkbook.active


    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import colors

    alertRule = ColorScaleRule(start_type='min', start_value=0, start_color=colors.WHITE, end_type='max', 
        end_value=100, end_color=colors.BLUE)

    outputMainSheet.conditional_formatting.add("F8:J250", alertRule)

    print("Conditional formatting applied successfully")

    global prtgDataDict
    prtgDataDict = {}

    ### [!] [Defining cliargs from returned param of cliArgumentParser() {To be stored globally}]
    #############
    global cliargs ### I know, globals are bad, but it saves a lot of typing in this situation
    global default_start
    global default_end
    global current_sys_datetime
    ### [!] [Assigning values to global vars]
    ############

    currentSystemDatetime = datetime.datetime.now()
    currentSystemDate = currentSystemDatetime.date()



    ## [user io] 
    cliargs = cliArgumentParser(currentSystemDatetime) # Parses CLI args

    ## [disk io] 
    xlsx_build() # Creates XLSX file and defines headers, column widths, row heights, etc. 
                 # Saves to 'test.xlsx', needs to be changed to reflect proper naming scheme
    print("XLSX file template built successfully")

    ## [api] 
    sensorKPIData = get_kpi_sensor_ids(cliargs.username, PRTG_PASSWORD, PRTG_HOSTNAME)
    print("PRTG Sensors retrieved successfully")
    
    kpi_seg_arr = []
    kpi_seg_arr.clear()
    s_count = 1

    prtgMainCall(sensorKPIData, PRTG_HOSTNAME, PRTG_PASSWORD, cliargs, kpi_seg_arr,s_count)
    

       

    outputWorkbook.save("test.xlsx")
