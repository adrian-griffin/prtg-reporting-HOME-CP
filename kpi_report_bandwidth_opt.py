### --------   PRTG-XLSX-Report-Generator.py
# -------------------------------------------------------------------------------
#
#   Pulls sensor & device data from PRTG API and neatly formats the data into a .xmlx (MS Excel) file
#       using python-openpyxl, json, csv, and pandas. 
#   
#       Average runtime: 274 seconds (4.5 - 5.0 minutes)
# -------------------------------------------------------------------------------
from time import time, time_ns
from tkinter import Frame
from turtle import width
import requests
import re
import math
import numpy
import getpass
import json

import datetime
import argparse
import openpyxl
import os

###########################################################
### [Most primary functions declared here]

### [Getting script's current working directory to be used later to ensure that 
#       no exceptions arise and for safer file IO]
#############



def cliArgumentParser(currentSystemDatetime):
    now = datetime.datetime.now()
    default_start = now - datetime.timedelta(days = 28)
    default_end = now - datetime.timedelta(days = 0)
    parser = argparse.ArgumentParser()
    parser.add_argument('--username', required=False, default="agriffin", ###### !! CHANGE FOR PROD !! #####
                    help='PRTG username for API call')
    parser.add_argument('--start', default=default_start.strftime('%Y-%m-%d'),
                    help='Historic data start date (yyyy-mm-dd)')
    parser.add_argument('--end', default=default_end.strftime('%Y-%m-%d'),
                    help='Historic data end date (yyyy-mm-dd)')
    parser.add_argument('--avgint', default="21600",
                    help='Averaging interval. Smaller numbers increase api call time!'
                        ' Default is 14400 seconds')
    parser.add_argument('--debug', action='store_true', dest='debug',
                    help='add additional debugging fields to output')
    parser.add_argument('--percentile', default="99",
                    help='set the percentile for reporting (default is 90)')
    parser.add_argument('--output',
                    help='specify output file directory (default is cwd)')
    parser.add_argument('--sensorid',
                    help='Pull data from only one sensorid')
    cliargs = parser.parse_args()
    return cliargs

def xlsx_build():

    now_datetime = datetime.datetime.now()
    now = now_datetime.date()

    t0BACK_headers = now
    t7BACK_headers = now - datetime.timedelta(days = 7)
    t14BACK_headers = now - datetime.timedelta(days = 14)
    t21BACK_headers = now - datetime.timedelta(days = 21)
    t28BACK_headers = now - datetime.timedelta(days = 28)

    t0BACK_headers = t0BACK_headers.strftime('%m/%d')
    t7BACK_headers = t7BACK_headers.strftime('%m/%d')
    t14BACK_headers = t14BACK_headers.strftime('%m/%d')
    t21BACK_headers = t21BACK_headers.strftime('%m/%d')
    t28BACK_headers = t28BACK_headers.strftime('%m/%d')

    sheetHeaders = ['Location','Highest Traffic (Mb/s)','Choke Point (Device)','Choke Point Throttle (Mb/s)','Circuit Max Limit (Mb/s)',
        'Circuit Utilization',
        f'Choke Utilization ({t0BACK_headers} - {t7BACK_headers})',
        f'Choke Utilization ({t7BACK_headers} -  {t14BACK_headers})',
        f'Choke Utilization ({t14BACK_headers} - {t21BACK_headers})',
        f'Choke Utilization ({t21BACK_headers} - {t28BACK_headers})',
        'Max Usage Plan','Notes','Action']

    coreUtilSummaryHeaders = ['Core Utilization Summary','Bandwidth (Mb/s)',
        f'Gross Utilization ({t0BACK_headers} - {t7BACK_headers})',
        f'Gross Utilization ({t7BACK_headers} - {t14BACK_headers})',
        f'Gross Utilization ({t14BACK_headers} - {t21BACK_headers})',
        f'Gross Utilization ({t21BACK_headers} - {t28BACK_headers})']

    global alphabetArray
    alphabetArray = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

    for letter in alphabetArray:
        outputMainSheet.column_dimensions[str(letter)].width = '16'
        outputSummarySheet.column_dimensions[str(letter)].width = '16'

    outputMainSheet.column_dimensions['A'].width = '28'
    outputSummarySheet.column_dimensions['A'].width = '28'
    outputMainSheet.column_dimensions['K'].width = '12'
    outputSummarySheet.column_dimensions['K'].width = '12'
    outputMainSheet.column_dimensions['F'].width = '12'
    outputSummarySheet.column_dimensions['F'].width = '16'
    outputMainSheet.column_dimensions['L'].width = '14'
    outputSummarySheet.column_dimensions['L'].width = '14'
    outputMainSheet.column_dimensions['M'].width = '14'
    outputSummarySheet.column_dimensions['M'].width = '14'
    outputMainSheet.column_dimensions['B'].width = '12'
    outputSummarySheet.column_dimensions['B'].width = '12'

    for i in range(0,len(sheetHeaders)):
        outputMainSheet[str(alphabetArray[i])+'1']=sheetHeaders[i]
        outputMainSheet[str(alphabetArray[i])+'1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)


    for i in range(0,len(coreUtilSummaryHeaders)):
        outputSummarySheet[str(alphabetArray[i])+'1']=coreUtilSummaryHeaders[i]
        outputSummarySheet[str(alphabetArray[i])+'1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    outputSummarySheet['A5']='Total: '

    #outputMainSheet['B2:M300'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    outputWorkbook.save("test.xlsx")

def writeToSheet(yIndex,xIndex,vIndex):
    if vIndex == '':
        if type(vIndex) == None:
            pass
        else:
            pass
    else:
        outputMainSheet.cell(row=int(yIndex),column=int(xIndex)).value=vIndex
    outputWorkbook.save("test.xlsx")

def writeToSummary(yIndex,xIndex,vIndex):
    if vIndex == '':
        pass
    else:
        outputSummarySheet.cell(row=int(yIndex),column=int(xIndex)).value=vIndex
    outputWorkbook.save("test.xlsx")

def get_kpi_sensor_ids(username, password, PRTG_HOSTNAME):
    """
    Returns a dict with the following format:
    {'prtg-version': '22.1.74.1869',
    'treesize': 2,
    'sensors': [{'objid': 14398,
    'objid_raw': 14398,
    'device': 'ACA Edge (160.3.214.2)',
    'device_raw': 'ACA Edge (160.3.214.2)'},
    """
   
    response = requests.get(
            f'https://{PRTG_HOSTNAME}/api/table.json?content=sensors&output=json'
            f'&columns=objid,device,tags&filter_tags=kpi_bandwidth'
            f'&username={username}&password={password}&sortby=device'
            )
    ### If 200 OK HTTP response is not seen, raise error and print cause to terminal
    if response.status_code == 200:
        response_tree = json.loads(response.text)
        print("Sensor data successfully queried from the API")
        if cliargs.sensorid:
            for i in response_tree.get('sensors'):
                if i['objid'] == int(cliargs.sensorid):
                    print(i)
                    return [i]
                    
        else:

            return response_tree.get('sensors')
    else:
        print("Error making API call to nanm.smartaira.net")
        print("HTTP response 200: OK was not received")
        print("Received response code: "+str(response.status_code))
        quit()

def normalize_traffic(data, label):
    """
    Takes an input of raw PRTG historic data and the label (ie 'Traffic In (speed)')
    and multiplies the speeds by 0.00008 to get values in mbits/sec ( ((8 / 10) / 100) / 1000)
    """
    traffic_list = []
    if data[label] != '':
        data[label] = data[label] * 0.000008
        traffic_list.append(data[label])

    else:
        print("Error Normalizing")
    return traffic_list

def extract_tags(sensor):
    """
    Takes a sensor dictionary and extacts a properties dictionary from the tags
    string returned by PRTG. The tags string will have a format something like this:
    'kpi_bandwidth kpi_seg=DIA kpi_choke=Circuit kpi_chokelimit=10000 kpi_cktmaxlimit=10000'
    
    """
    target_tags = ['kpi_seg', 'kpi_choke', 'kpi_cktmaxlimit', 'kpi_siteid', 'edge']
    def filter_tags(tags_list, property_string):
        """
        Takes a list of tags and splits them into a properties dict:
        ['kpi_choke=Circuit', kpi_chokelimit=10000'] becomes
        {'kpi_choke':'Circuit', 'kpi_chokelimit':'10000'}
        
        This function exists because we might have tags like kpi_choke and kpi_chokelimit
        that will both be found by the _in_ function
        """
        properties = {}
        property_list = list(filter(lambda a: property_string in a, tags_list))
        for property in property_list:
            try:
                key, value = property.split('=')
            except ValueError:
                key = 'edge'
                value = 'True'
            properties.update({key:value})    
            
        return properties

    device_properties = {}
    tag_string = sensor['tags'].split()

    for tag in target_tags:
        device_properties.update(filter_tags(tag_string, tag))

    return device_properties

def buildComps(loc_index,FrameWindow,datablock,historicResponseData,sensor,sensor_index):
    try:
        prtgDataDict[f'{FrameWindow} Weeks Back'] = {
      "FrameWindow": f'{FrameWindow} Weeks Back',
      f'{FrameWindow} Weeks Back': [
    {
      "datetime": datablock['datetime'],
      "Traffic Total (volume)": datablock['Traffic Total (volume)'],
      "Traffic Total (speed)": datablock['Traffic Total (speed)'],
      "Traffic In (volume)": datablock['Traffic In (volume)'],
      "Traffic In (speed)": datablock['Traffic In (speed)'],
      "Traffic Out (volume)": datablock['Traffic Out (volume)'],
      "Traffic Out (speed)": datablock['Traffic Out (speed)'],
      "Downtime": datablock['Downtime'],
      "coverage": datablock['coverage']
    }]}
        if FrameWindow == 0:
            prtgMainParse(FrameWindow,sensor,kpi_seg_arr,s_count,loc_index,datablock)
        else: 
            prtgExtendHistParse(FrameWindow,sensor,kpi_seg_arr,s_count,loc_index,datablock)
    except KeyError:
        pass

def storeAPIResponse(loc_index,sensordata,historicResponseData,sensor,sensor_index):
    #    prtgDataDict[str(sensor['device'])]['objid'] = sensor['objid']
    i = 0
    while i < int(historicResponseData['treesize']):
        #dateBasedSectioning(sensordata, historicResponseData, sensor, sensor_index)
        indexHistoryDictionary = historicResponseData['histdata'][i]
        prtg_formatted_datetime = indexHistoryDictionary['datetime']
        try:
            regex_datetimesrc = re.search(r'\d{1}/\d{2}/\d{4}',prtg_formatted_datetime)
        except AttributeError:
            pass 
        if regex_datetimesrc:
            py_formatted_datetime = datetime.datetime.strptime(regex_datetimesrc.group(), '%m/%d/%Y')
        else: 
            try:
                regex_datetimesrc = re.search(r'\d{2}/\d{2}/\d{4}',prtg_formatted_datetime)
            except AttributeError:
                print("No date found in 'datetime' object!")
                pass
            if regex_datetimesrc:
                py_formatted_datetime = datetime.datetime.strptime(regex_datetimesrc.group(), '%m/%d/%Y')

        py_formatted_date = py_formatted_datetime.date()

        now_datetime = datetime.datetime.now()
        now = now_datetime.date()

        if py_formatted_date > now:
            print("Received PRTG data has out of range timestamp! (In the future)")
        elif py_formatted_date < (now - datetime.timedelta(days = 28)):
            print("Received PRTG data has out of range timestamp! (Too far back)")

        elif py_formatted_date > (now - datetime.timedelta(days = 7)):
            historicResponseData['histdata'][i]['datetime'] = str(py_formatted_date)
            buildComps(loc_index,0,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)
            
        elif py_formatted_date < (now - datetime.timedelta(days = 21)):
            historicResponseData['histdata'][i]['datetime'] = str(py_formatted_date)
            buildComps(loc_index,1,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        elif py_formatted_date > (now - datetime.timedelta(days = 21)) and py_formatted_date < (now - datetime.timedelta(days = 14)):
            historicResponseData['histdata'][i]['datetime'] = str(py_formatted_date)
            buildComps(loc_index,2,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        elif py_formatted_date < (now - datetime.timedelta(days = 7)) and py_formatted_date > (now - datetime.timedelta(days = 14)):
            historicResponseData['histdata'][i]['datetime'] = str(py_formatted_date)
            buildComps(loc_index,3,historicResponseData['histdata'][i],historicResponseData,sensor,sensor_index)

        else: 
            pass


        i += 1
    '''
    for key,value in sensor.items():
        prtgDataDict[str(sensor['tags'])] = {}
        sensor_tags = sensor['tags']
        prtgDataDict[str(sensor['tags'])]['bandwidthsensor kpi_bandwidth kpi_choke=Circuit kpi_chokelimit=1000 kpi_cktmaxlimit=1000 kpi_seg=DIA kpi_siteid=94th#Aero router snmptrafficsensor router']
    '''

def prtgExtendHistParse(FrameWindow,sensor,kpi_seg_arr,s_count,i,datablock):

    data = prtgDataDict[f'{FrameWindow} Weeks Back'][f'{FrameWindow} Weeks Back'][0]
    sensorTagData = extract_tags(sensor)
    properties = extract_tags(sensor)
    traffic_in = normalize_traffic(data, 'Traffic In (speed)')
    traffic_out = normalize_traffic(data, 'Traffic Out (speed)')
    device_name = sensor['device'].split(' (')[0]

    ### [dec] - [CIRCUIT MAX LIMIT (Mb/s)]
    #############
    if sensorTagData.get('kpi_cktmaxlimit'):
        writeToSheet(i,5,float(sensorTagData['kpi_cktmaxlimit']))
    else:
        writeToSheet(i,5,'NA')


    ### [dec] - [MAX TRAFFIC (Mb/s)]
    #############
    max_traffic = 0
    if properties.get('kpi_trafficdirection') == 'up':
        if traffic_out == []:
            max_traffic_san = 0
        else:
            max_traffic = math.ceil(numpy.percentile(traffic_out, int(cliargs.percentile)))
            max_traffic_san = max_traffic
    else:
        if traffic_in == []:
            max_traffic_san = 0
        else:
            max_traffic = math.ceil(numpy.percentile(traffic_in, int(cliargs.percentile)))
            max_traffic_san = max_traffic


    if "Core" in sensorTagData.get('kpi_seg'):
        s_count = len(kpi_seg_arr)+1
        if sensorTagData.get('kpi_seg') in kpi_seg_arr:
            pass
        else:
            writeToSummary(1+s_count,3+FrameWindow,(float(max_traffic)/float(sensorTagData.get('kpi_cktmaxlimit'))))
            outputSummarySheet.cell(row=int(1+s_count),column=3+FrameWindow).style='Percent'
            

    ### [dec] - [CHOKE POINT UTILIZATION (%)]

    if properties.get('kpi_chokelimit'):
        outputMainSheet.cell(row=int(i),column=7+FrameWindow).value=(float(max_traffic) / float(properties['kpi_chokelimit']))
        outputMainSheet.cell(row=int(i),column=7+FrameWindow).style='Percent'
        outputSummarySheet.cell(row=int(1+s_count),column=3).style='Percent'
    else:
        outputMainSheet.cell(row=int(i),column=7+FrameWindow).value='NA'

def prtgMainParse(FrameWindow,sensor,kpi_seg_arr,s_count,i,datablock):
    
    mainResponseData = prtgDataDict[f'{FrameWindow} Weeks Back'][f'{FrameWindow} Weeks Back'][0]
    sensorTagData = extract_tags(sensor)
    trafficInbound = normalize_traffic(mainResponseData, 'Traffic In (speed)')
    trafficOutbound = normalize_traffic(mainResponseData, 'Traffic Out (speed)')
    deviceName = sensor['device'].split(' (')[0]

    ### [dec] - [LOCATION (Location)]
    #############
    if sensorTagData.get('kpi_siteid'):
        writeToSheet(i,1,re.sub("#", " ", sensorTagData['kpi_siteid']))
    else:
        writeToSheet(i,1,'NA')

    if cliargs.debug:
        # Device name (debug)
        print(deviceName)
        # Device id (debug)
        print(sensor['objid'])

    ### [dec] - [MAX TRAFFIC (Mb/s)]
    #############
    max_traffic = 0
    if sensorTagData.get('kpi_trafficdirection') == 'up':
        if trafficOutbound == []:
            writeToSheet(i,2,'NA')
        else:
            max_traffic = math.ceil(numpy.percentile(trafficOutbound, int(cliargs.percentile)))
            writeToSheet(i,2,max_traffic)
    else:
        if trafficInbound == []:
            writeToSheet(i,2,'NA')
        else:
            max_traffic = math.ceil(numpy.percentile(trafficInbound, int(cliargs.percentile)))
            writeToSheet(i,2,max_traffic)

    ### [dec] - [CHOKE POINT (Device)]
    #############
    if sensorTagData.get('kpi_choke'):
        writeToSheet(i,3,sensorTagData['kpi_choke'])
    else:
        outputMainSheet.cell(row=int(i),column=3).value='NA'
        writeToSheet(i,3,'NA')

    ### [dec] - [CHOKE POINT LIMIT (Mb/s)]
    #############
    if sensorTagData.get('kpi_chokelimit'):
        writeToSheet(i,4,float(sensorTagData['kpi_chokelimit']))
        
    else:
        writeToSheet(i,4,'NA')
    


    ### [dec] - [CIRCUIT MAX LIMIT (Mb/s)]
    #############
    if sensorTagData.get('kpi_cktmaxlimit'):
        writeToSheet(i,5,float(sensorTagData['kpi_cktmaxlimit']))
    else:
        writeToSheet(i,5,'NA')

    ### [dec] - [CIRCUIT UTILIZATION (%)]
    #############
    if sensorTagData.get('kpi_cktmaxlimit'):
        writeToSheet(i,6,(float(max_traffic) / float(sensorTagData['kpi_cktmaxlimit'])))
        outputMainSheet.cell(row=int(i),column=int(6)).style='Percent'
    else:
        writeToSheet(i,6,'NA')


    ### [dec] - [CHOKE POINT UTILIZATION (%) 1]
    #############
    if sensorTagData.get('kpi_chokelimit'):
        writeToSheet(i,7,(float(max_traffic) / float(sensorTagData['kpi_chokelimit'])))
        outputMainSheet.cell(row=int(i),column=int(7)).style='Percent'
    else:
        writeToSheet(i,7,'NA') 


    if sensorTagData.get('edge'):
        outputMainSheet.cell(row=int(i),column=int(7)).style='Percent'
        
    else:
        pass    

    ### [dec] - [EDGE BOOLEAN]
    #############
    if sensorTagData.get('edge'):
        edge_count = 1
        while edge_count <= 6:
            cellToColourBlue_letter = alphabetArray[edge_count-1]
            cellToColourBlue_number = i
            outputMainSheet[f'{cellToColourBlue_letter}{cellToColourBlue_number}'].fill = openpyxl.styles.PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type = "solid")
            edge_count += 1

    else:
        pass


    if "Core" in sensorTagData.get('kpi_seg'):
        s_count = len(kpi_seg_arr)+1
        if sensorTagData.get('kpi_seg') in kpi_seg_arr:
            pass
        else:
            writeToSummary(1+s_count,1,sensorTagData.get('kpi_seg'))
            writeToSummary(1+s_count,2,max_traffic)
            writeToSummary(1+s_count,3,(float(max_traffic)/float(sensorTagData.get('kpi_cktmaxlimit'))))
            outputSummarySheet.cell(row=int(1+s_count),column=3).style='Percent'
            s_count += 1
            kpi_seg_arr.append(sensorTagData.get('kpi_seg'))
            print("")
            print("------------------")
            print("CORE Device Added!")
            print("------------------")
            print("")

def prtgMainCall(sensordata,PRTG_HOSTNAME,PRTG_PASSWORD,cliargs,kpi_seg_arr,s_count):
    loc_index = 2
    sensor_index = 0
    print("Writing API query results to cache")
    for sensor in sensordata:
        response = api_session.get(
            f'https://{PRTG_HOSTNAME}/api/historicdata.json?id={sensor["objid"]}'
            f'&avg={cliargs.avgint}&sdate={cliargs.start}-00-00&edate={cliargs.end}-0-0'
            f'&usecaption=1'
            f'&username={cliargs.username}&password={PRTG_PASSWORD}'
            )
        if response.status_code == 200:
            response_j = response.json()

            storeAPIResponse(loc_index,sensordata,response_j,sensor,sensor_index)
            outputSummarySheet['B5']='=SUM(B2:B4)'
            outputSummarySheet['C5']='=SUM(C2:C4)'
            outputSummarySheet['D5']='=SUM(D2:D4)'
            outputSummarySheet['E5']='=SUM(E2:E4)'
            outputSummarySheet['F5']='=SUM(F2:F4)'
            

        else:
            print("Error making 'main' API call to nanm.smartaira.net (PRTG)")
            print("HTTP response 200: OK was not received")
            print("Received response code: "+str(response.status_code))
            exit(1)

        loc_index += 1
        sensor_index += 1
### [PRTG API call and assigning data to "sensors" var]
#############

if __name__ == '__main__':
    # IP hostname & requests.session   : Time to run: 941.9562714099884 s [15 min]
    # IP hostname & requests.get       : Time to run: 864.3155009746552 s [14 min]
    # URL hosthame & requests.get      : Time to run: 845.8513550758362 s [14 min]
    # URL hostname & requests.session  : Time to run: 936.6703135967255 s [15 min]
    import time
    beginTime = time.time()


    PRTG_PASSWORD = "M9y%23asABUx9svvs"  ###### !! CHANGE FOR PROD !! ##### ----------
    PRTG_HOSTNAME = 'nanm.smartaira.net'   ###!! Static, domain/URL to PRTG server
    #PRTG_HOSTNAME = '64.71.154.163'
    #PRTG_PASSWORD = getpass.getpass('Password: ')

    outputWorkbook = openpyxl.Workbook()

    sheet1_to_be_del = outputWorkbook.get_sheet_by_name('Sheet')
    outputWorkbook.remove_sheet(sheet1_to_be_del)

    outputMainSheet = outputWorkbook.create_sheet("Property Bandwidths")
    outputSummarySheet = outputWorkbook.create_sheet("Summaries")

    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import colors

    alertRule = ColorScaleRule(start_type='min', start_value=0, start_color=colors.WHITE, end_type='max', 
        end_value=100, end_color='F8696B')

    outputMainSheet.conditional_formatting.add("F2:J250", alertRule)
    outputSummarySheet.conditional_formatting.add("C2:F40", alertRule)

    global prtgDataDict
    prtgDataDict = {}

    ### [!] [Defining cliargs from returned param of cliArgumentParser() {To be stored globally}]
    #############
    global cliargs ### I know, globals are bad, but it saves a lot of typing in this situation
    global default_start
    global default_end
    ### [!] [Assigning values to global vars]
    ############

    currentSystemDatetime = datetime.datetime.now()
    currentSystemDate = currentSystemDatetime.date()


    api_session = requests.Session()

    ## [user io] 
    cliargs = cliArgumentParser(currentSystemDatetime) # Parses CLI args

    ## [disk io] 
    xlsx_build() # Creates XLSX file and defines headers, column widths, row heights, etc. 
                 # Saves to 'test.xlsx', needs to be changed to reflect proper naming scheme
    print("XLSX file template built successfully")

    ## [api] 
    sensorKPIData = get_kpi_sensor_ids(cliargs.username, PRTG_PASSWORD, PRTG_HOSTNAME)
    print("PRTG Sensors retrieved successfully")
    
    global kpi_seg_arr
    global s_count
    kpi_seg_arr = []
    kpi_seg_arr.clear()
    s_count = 1

    prtgMainCall(sensorKPIData, PRTG_HOSTNAME, PRTG_PASSWORD, cliargs, kpi_seg_arr,s_count)
    

    finalTime = time.time()
    timeToRun = str(finalTime - beginTime)
    timeToRunMinutes = str(int(float(timeToRun)/float(60)))
    print("")
    print("")
    print("##############")
    print(f'Time to run: {timeToRun} s [{timeToRunMinutes} min]')
    print("##############")
    print("")
    print("")
    outputWorkbook.save("test.xlsx")
